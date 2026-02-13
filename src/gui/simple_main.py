import os
import csv
import sys
import traceback
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import xlrd  # 用于读取Excel文件
from openpyxl import Workbook  # 替换xlwt，使用openpyxl
from openpyxl import load_workbook  # 用于读取xlsx格式
from collections import defaultdict

from src.core.admission import (
    ADJUST_SUFFIX,
    INVALID_CHOICE_LABEL,
    UNASSIGNED_LABEL,
    assign_admissions,
)
from src.core.preferences import PREFERENCE_MAPPING

# 设置日志
def setup_logging():
    log_dir = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'app.log')
    
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

def get_resource_path(relative_path):
    """获取资源文件的绝对路径"""
    try:
        # PyInstaller创建临时文件夹,将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except Exception:
        # 如果不是打包的情况,则使用当前文件的目录
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    
    return os.path.join(base_path, relative_path)

class SimpleMajorAdmissionApp:
    def __init__(self, root):
        try:
            self.root = root
            self.root.title("本科生专业方向录取软件 V1.0")
            self.root.geometry("800x700")  # 增加窗口高度以适应LOGO
            
            # 添加异常处理
            self.root.report_callback_exception = self.handle_exception
            
            # Initialize data
            self.student_data = []
            self.major_quotas = {
                "电子信息工程": tk.IntVar(value=0),
                "通信工程": tk.IntVar(value=0),
                "电磁场与无线技术": tk.IntVar(value=0)
            }
            
            self.preference_mapping = PREFERENCE_MAPPING
            
            self.init_ui()
            
            # 创建菜单栏
            self.create_menu()
            
        except Exception as e:
            logging.error(f"初始化失败: {str(e)}")
            logging.error(traceback.format_exc())
            messagebox.showerror("错误", f"程序初始化失败：{str(e)}\n请查看日志文件了解详情。")
    
    def create_menu(self):
        """创建菜单栏"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用说明", command=self.show_help)
        help_menu.add_command(label="关于", command=self.show_about)
    
    def show_help(self):
        """显示使用说明对话框"""
        help_window = tk.Toplevel(self.root)
        help_window.title("使用说明")
        help_window.geometry("600x500")
        help_window.resizable(False, False)
        
        # 设置模态
        help_window.transient(self.root)
        help_window.grab_set()
        
        # 创建主框架，添加滚动条
        main_frame = ttk.Frame(help_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 添加使用说明
        help_text = """使用说明

一、软件说明
本系统用于处理本科生专业方向录取工作，支持按照学生成绩和志愿自动分配专业方向。

系统特点：
• 全自动处理录取流程
• 支持多种文件格式（Excel、CSV）
• 可视化操作界面
• 自动生成统计报告
• 支持导出录取结果

二、使用步骤
1. 设置专业名额
   - 在界面上方设置各专业录取名额
   - 电子信息工程
   - 通信工程
   - 电磁场与无线技术

2. 导入学生数据
   - 点击"导入学生志愿"按钮
   - 选择Excel文件（.xlsx/.xls）或CSV文件
   - 确保文件格式正确（包含必要列）

3. 处理录取
   - 点击"处理录取"按钮
   - 系统将自动按照以下规则处理：
     * 按成绩从高到低排序
     * 优先满足第一志愿
     * 未被录取者依次考虑第二、第三志愿
     * 仍未被录取者进入调剂

4. 查看/导出结果
   - 界面下方表格实时显示录取结果
   - 点击"导出录取结果"保存为Excel文件

三、志愿代码说明
A：电子信息工程 > 通信工程 > 电磁场与无线技术
B：电子信息工程 > 电磁场与无线技术 > 通信工程
C：电磁场与无线技术 > 电子信息工程 > 通信工程
D：电磁场与无线技术 > 通信工程 > 电子信息工程
E：通信工程 > 电子信息工程 > 电磁场与无线技术
F：通信工程 > 电磁场与无线技术 > 电子信息工程

四、注意事项
1. 数据文件要求
   - 表头必须包含：序号、学号、姓名、分数、志愿选择
   - 志愿选择使用A-F表示不同的志愿组合
   - 分数必须为数字格式

2. 使用建议
   - 正式使用前先用测试数据运行
   - 及时导出并保存录取结果
   - 定期备份重要数据文件"""
        
        # 使用Text控件显示帮助文本，支持选择复制
        help_text_widget = tk.Text(scrollable_frame, wrap=tk.WORD, width=60, height=20)
        help_text_widget.insert('1.0', help_text)
        help_text_widget.configure(state='disabled')  # 设置为只读
        help_text_widget.pack(pady=5, fill=tk.BOTH, expand=True)
        
        # 布局滚动条和画布
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 确定按钮
        ttk.Button(help_window, text="确定", command=help_window.destroy).pack(pady=10)
        
        # 居中显示
        help_window.update_idletasks()
        width = help_window.winfo_width()
        height = help_window.winfo_height()
        x = (help_window.winfo_screenwidth() // 2) - (width // 2)
        y = (help_window.winfo_screenheight() // 2) - (height // 2)
        help_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        
        # 绑定鼠标滚轮事件
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except Exception as e:
                pass  # 忽略任何错误，因为窗口可能已经关闭
        
        # 只绑定到canvas上，而不是全局绑定
        canvas.bind("<MouseWheel>", _on_mousewheel)
        
        # 当关闭窗口时清理
        def _on_closing():
            try:
                canvas.unbind("<MouseWheel>")
                help_window.destroy()
            except Exception as e:
                pass  # 忽略任何错误，因为窗口可能已经关闭
        
        help_window.protocol("WM_DELETE_WINDOW", _on_closing)

    def show_about(self):
        """显示关于对话框"""
        about_window = tk.Toplevel(self.root)
        about_window.title("关于")
        about_window.geometry("400x300")
        about_window.resizable(False, False)
        
        # 设置模态
        about_window.transient(self.root)
        about_window.grab_set()
        
        # 创建主框架
        main_frame = ttk.Frame(about_window)
        main_frame.pack(expand=True)
        
        # 添加Logo
        try:
            logo_path = get_resource_path(os.path.join('resources', 'logo.png'))
            if not os.path.exists(logo_path):
                logo_path = get_resource_path(os.path.join('resources', 'logo.ico'))
            
            if os.path.exists(logo_path):
                logo_img = Image.open(logo_path)
                # 调整大小为100x100
                logo_img = logo_img.resize((100, 100), Image.Resampling.LANCZOS)
                logo_photo = ImageTk.PhotoImage(logo_img)
                logo_label = ttk.Label(main_frame, image=logo_photo)
                logo_label.image = logo_photo
                logo_label.pack(pady=10)
        except Exception as e:
            logging.warning(f"加载关于窗口Logo失败: {str(e)}")
        
        # 添加基本信息
        ttk.Label(main_frame, text="电子信息与通信学院", font=("Arial", 14, "bold")).pack(pady=5)
        ttk.Label(main_frame, text="本科生专业方向录取系统", font=("Arial", 12)).pack()
        ttk.Label(main_frame, text="版本：V1.0", font=("Arial", 10)).pack(pady=5)
        ttk.Label(main_frame, text="© 2024 电子信息与通信学院", font=("Arial", 10)).pack(pady=5)
        ttk.Label(main_frame, text="作者：lucasxuchang@icloud.com", font=("Arial", 10)).pack(pady=5)
        
        
        
        # 居中显示
        about_window.update_idletasks()
        width = about_window.winfo_width()
        height = about_window.winfo_height()
        x = (about_window.winfo_screenwidth() // 2) - (width // 2)
        y = (about_window.winfo_screenheight() // 2) - (height // 2)
        about_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    def handle_exception(self, exc_type, exc_value, exc_traceback):
        """处理未捕获的异常"""
        error_msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        logging.error(f"未捕获的异常:\n{error_msg}")
        messagebox.showerror("错误", f"发生错误：{str(exc_value)}\n请查看日志文件了解详情。")

    def init_ui(self):
        try:
            # Create main frame
            main_frame = ttk.Frame(self.root, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Major quotas input section
            quotas_frame = ttk.LabelFrame(main_frame, text="专业录取名额设置", padding="10")
            quotas_frame.pack(fill=tk.X, pady=10)
            
            quotas_inner_frame = ttk.Frame(quotas_frame)
            quotas_inner_frame.pack(fill=tk.X)
            
            for i, (major, var) in enumerate(self.major_quotas.items()):
                major_frame = ttk.Frame(quotas_inner_frame)
                major_frame.pack(side=tk.LEFT, padx=10, expand=True)
                
                ttk.Label(major_frame, text=major).pack()
                spin_box = ttk.Spinbox(
                    major_frame, 
                    from_=0, 
                    to=1000, 
                    textvariable=var,
                    width=10
                )
                spin_box.pack(pady=5)
            
            # File operations section
            file_operations_frame = ttk.Frame(main_frame)
            file_operations_frame.pack(fill=tk.X, pady=10)
            
            import_btn = ttk.Button(file_operations_frame, text="导入学生志愿", command=self.import_student_data)
            import_btn.pack(side=tk.LEFT, padx=5)
            
            process_btn = ttk.Button(file_operations_frame, text="处理录取", command=self.process_admissions)
            process_btn.pack(side=tk.LEFT, padx=5)
            
            export_btn = ttk.Button(file_operations_frame, text="导出录取结果", command=self.export_results)
            export_btn.pack(side=tk.LEFT, padx=5)
            
            # Results table
            table_frame = ttk.LabelFrame(main_frame, text="录取结果", padding="10")
            table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Create treeview for results
            self.results_tree = ttk.Treeview(table_frame, columns=("序号", "学号", "姓名", "分数", "志愿选择", "录取专业"), show="headings")
            
            # Define columns
            self.results_tree.heading("序号", text="序号")
            self.results_tree.heading("学号", text="学号")
            self.results_tree.heading("姓名", text="姓名")
            self.results_tree.heading("分数", text="分数")
            self.results_tree.heading("志愿选择", text="志愿选择")
            self.results_tree.heading("录取专业", text="录取专业")
            
            # Set column widths
            self.results_tree.column("序号", width=50)
            self.results_tree.column("学号", width=100)
            self.results_tree.column("姓名", width=100)
            self.results_tree.column("分数", width=80)
            self.results_tree.column("志愿选择", width=80)
            self.results_tree.column("录取专业", width=150)
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
            self.results_tree.configure(yscroll=scrollbar.set)
            
            # Pack treeview and scrollbar
            self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        except Exception as e:
            logging.error(f"初始化UI失败: {str(e)}")
            logging.error(traceback.format_exc())
            messagebox.showerror("错误", f"初始化UI失败：{str(e)}\n请查看日志文件了解详情。")
    
    def import_student_data(self):
        try:
            file_name = filedialog.askopenfilename(
                title="选择学生志愿文件",
                filetypes=[("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
            
            if file_name:
                self.student_data = []
                
                if file_name.endswith('.csv'):
                    # 使用csv模块读取csv文件
                    with open(file_name, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            student = {
                                '序号': row['序号'],
                                '排名': row.get('排名', row['序号']),
                                '学号': row['学号'],
                                '姓名': row['姓名'],
                                '分数': float(row['分数']),
                                '志愿选择': str(row['志愿选择']).upper(),  # 转换为大写
                                '专业': row['专业']
                            }
                            self.student_data.append(student)
                
                elif file_name.endswith('.xlsx'):
                    # 使用openpyxl读取xlsx文件
                    wb = load_workbook(file_name)
                    sheet = wb.active
                    
                    # 读取数据
                    for row in sheet.iter_rows(min_row=2):
                        student = {
                            '序号': row[0].value,
                            '排名': row[0].value,
                            '学号': row[1].value,
                            '姓名': row[2].value,
                            '分数': float(row[4].value),
                            '志愿选择': str(row[6].value).upper(),  # 转换为大写
                            '专业': row[7].value
                        }
                        self.student_data.append(student)
                else:
                    # 使用xlrd读取xls文件
                    workbook = xlrd.open_workbook(file_name)
                    sheet = workbook.sheet_by_index(0)
                    
                    # 读取数据
                    for row_idx in range(1, sheet.nrows):
                        student = {
                            '序号': sheet.cell_value(row_idx, 0),
                            '排名': sheet.cell_value(row_idx, 0),
                            '学号': sheet.cell_value(row_idx, 1),
                            '姓名': sheet.cell_value(row_idx, 2),
                            '分数': float(sheet.cell_value(row_idx, 4)),
                            '志愿选择': str(sheet.cell_value(row_idx, 6)).upper(),  # 转换为大写
                            '专业': sheet.cell_value(row_idx, 7)
                        }
                        self.student_data.append(student)
                
                self.update_results_table()
                messagebox.showinfo("成功", f"成功导入 {len(self.student_data)} 条学生数据")
        except Exception as e:
            messagebox.showerror("错误", f"导入文件时发生错误：{str(e)}")
            logging.error(f"导入文件时发生错误: {str(e)}")
            logging.error(traceback.format_exc())
    
    def process_admissions(self):
        if not self.student_data:
            messagebox.showwarning("警告", "请先导入学生数据")
            return
        
        try:
            # 获取当前名额
            quotas = {major: var.get() for major, var in self.major_quotas.items()}
            
            # 检查是否所有专业都设置了名额
            if all(quota == 0 for quota in quotas.values()):
                messagebox.showwarning("警告", "请先设置专业录取名额")
                return
                
            result = assign_admissions(
                self.student_data,
                quotas,
                self.preference_mapping,
                score_key=("排名" if all("排名" in s for s in self.student_data) else "分数"),
                sort_desc=False if all("排名" in s for s in self.student_data) else True,
                choice_key="志愿选择",
                assigned_key="录取专业",
            )

            # Keep UI state consistent with assigned/sorted order.
            self.student_data = result.students
            remaining_quotas = result.remaining_quotas
            
            self.update_results_table()
            
            # 统计录取信息
            total_students = len(self.student_data)
            invalid_count = sum(
                1 for s in self.student_data if s.get("录取专业", "") == INVALID_CHOICE_LABEL
            )
            unassigned_count = sum(
                1 for s in self.student_data if s.get("录取专业", "") == UNASSIGNED_LABEL
            )
            admitted_count = total_students - invalid_count - unassigned_count
            not_admitted_count = invalid_count + unassigned_count
            
            stats = {
                '电子信息工程': {'total': 0, 'adjust': 0},
                '通信工程': {'total': 0, 'adjust': 0},
                '电磁场与无线技术': {'total': 0, 'adjust': 0},
                '未分配': 0,
                '无效志愿': 0,
            }
            
            for student in self.student_data:
                major = student.get('录取专业', '')
                if major == UNASSIGNED_LABEL:
                    stats['未分配'] += 1
                    continue
                if major == INVALID_CHOICE_LABEL:
                    stats['无效志愿'] += 1
                    continue
                else:
                    base_major = major
                    is_adjust = False
                    if base_major.endswith(ADJUST_SUFFIX):
                        base_major = base_major[: -len(ADJUST_SUFFIX)]
                        is_adjust = True
                    if base_major in stats:
                        stats[base_major]['total'] += 1
                        if is_adjust:
                            stats[base_major]['adjust'] += 1
            
            # 生成详细的统计信息
            result_msg = "录取完成！\n\n"
            result_msg += f"总人数：{total_students}人\n"
            result_msg += f"已录取：{admitted_count}人\n"
            result_msg += f"未录取：{not_admitted_count}人\n\n"
            result_msg += "各专业录取情况：\n"
            
            for major, data in stats.items():
                if major != '未分配':
                    total = data['total']
                    adjust = data['adjust']
                    normal = total - adjust
                    result_msg += f"\n{major}：\n"
                    result_msg += f"  - 总计：{total}人\n"
                    result_msg += f"  - 正常录取：{normal}人\n"
                    result_msg += f"  - 调剂录取：{adjust}人\n"
                    result_msg += f"  - 剩余名额：{remaining_quotas[major]}人\n"
            
            result_msg += f"\n未分配人数：{stats['未分配']}人"
            result_msg += f"\n无效志愿人数：{stats['无效志愿']}人"
            
            messagebox.showinfo("录取完成", result_msg)
            
        except Exception as e:
            messagebox.showerror("错误", f"处理录取时发生错误：{str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def export_results(self):
        if not self.student_data:
            messagebox.showwarning("警告", "请先导入学生数据")
            return
            
        try:
            file_name = filedialog.asksaveasfilename(
                title="保存录取结果",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )
            
            if file_name:
                # 创建新的工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = '录取结果'
                
                # 写入表头
                headers = ['序号', '学号', '姓名', '分数', '志愿选择', '录取专业']
                ws.append(headers)
                
                # 写入数据
                for student in self.student_data:
                    ws.append([
                        student['序号'],
                        student['学号'],
                        student['姓名'],
                        student['分数'],
                        student['志愿选择'],
                        student.get('录取专业', '')
                    ])
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # 保存文件
                wb.save(file_name)
                
                messagebox.showinfo("成功", "录取结果已成功导出")
                
                # 询问是否打开文件
                if messagebox.askyesno("确认", "是否立即打开导出的文件？"):
                    os.startfile(file_name)
                
        except Exception as e:
            messagebox.showerror("错误", f"导出文件时发生错误：{str(e)}")
            logging.error(f"导出文件时发生错误: {str(e)}")
            logging.error(traceback.format_exc())
    
    def update_results_table(self):
        # 清除现有项目
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # 添加数据到树形视图
        for student in self.student_data:
            self.results_tree.insert(
                "", 
                tk.END, 
                values=(
                    student['序号'],
                    student['学号'],
                    student['姓名'],
                    student['分数'],
                    student['志愿选择'],
                    student.get('录取专业', '')
                )
            )

def main():
    try:
        # 设置日志
        setup_logging()
        logging.info("程序启动")
        
        # 创建主窗口
        root = tk.Tk()
        
        # 设置窗口图标
        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            icon_path = os.path.join(base_path, "..", "..", "resources", "logo.ico")
            if os.path.exists(icon_path):
                root.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"设置窗口图标失败: {str(e)}")
        
        app = SimpleMajorAdmissionApp(root)
        root.mainloop()
        
    except Exception as e:
        logging.error(f"程序运行失败: {str(e)}")
        logging.error(traceback.format_exc())
        messagebox.showerror("错误", f"程序运行失败：{str(e)}\n请查看日志文件了解详情。")
        sys.exit(1)

if __name__ == "__main__":
    main() 
